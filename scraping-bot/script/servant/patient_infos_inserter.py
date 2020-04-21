import os
import uuid
import settings as env
import openpyxl as px
from servant import file_downloader


def __convert_to_iso(date):
    # エクセルDatetimeをISOのDatetimeに変換する
    return date.isoformat(timespec='seconds')


def serve(db):
    # 既にエクセルファイルがある場合一旦、既存のファイルを破棄
    if os.path.isfile(env.DL_FILE_NAME):
        os.remove(env.DL_FILE_NAME)

    file_downloader.serve(env.DL_FILE, env.DL_FILE_NAME)

    # エクセルファイルを開いてシートを読みこむ
    wb = px.load_workbook(env.DL_FILE_NAME, data_only=True)
    sheet = wb["yousei"]

    for irow in range(2, sheet.max_row+1):
        date = __convert_to_iso(sheet.cell(row=irow, column=1).value)
        pcr_total = sheet.cell(row=irow, column=3).value
        positive_up_total = sheet.cell(row=irow, column=4).value
        positive_up_daily = (sheet.cell(row=irow, column=4).value if irow == 2 else
                             sheet.cell(row=irow, column=4).value - sheet.cell(row=irow-1, column=4).value)
        positive = sheet.cell(row=irow, column=5).value
        positive_yesterday = 0 if irow == 2 else sheet.cell(row=irow-1, column=5).value
        not_severe = sheet.cell(row=irow, column=6).value
        severe = sheet.cell(row=irow, column=7).value
        death_total = sheet.cell(row=irow, column=8).value
        death_daily = (sheet.cell(row=irow, column=8).value if irow == 2 else
                       sheet.cell(row=irow, column=8).value - sheet.cell(row=irow-1, column=8).value)
        discharge_total = sheet.cell(row=irow, column=9).value
        discharge_daily = (sheet.cell(row=irow, column=9).value if irow == 2 else
                           sheet.cell(row=irow, column=9).value - sheet.cell(row=irow - 1, column=9).value)

        db.execute("SELECT EXISTS (SELECT * FROM patient_infos WHERE date = %s);", (date,))
        (is_exist_recode,) = db.fetchone()

        if not is_exist_recode:
            _id = uuid.uuid4()

            db.execute("""
            INSERT
            INTO patient_infos (
                id,
                date,
                pcr_total,
                positive_up_total,
                positive_up_daily,
                positive,
                positive_yesterday,
                not_severe,
                severe,
                death_total,
                death_daily,
                discharge_total,
                discharge_daily
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
            """, (
                str(_id),
                date,
                pcr_total,
                positive_up_total,
                positive_up_daily,
                positive,
                positive_yesterday,
                not_severe,
                severe,
                death_total,
                death_daily,
                discharge_total,
                discharge_daily
            ))
        else:

            db.execute("""
            UPDATE patient_infos
            SET pcr_total = %s,
                positive_up_total = %s,
                positive_up_daily = %s,
                positive = %s,
                positive_yesterday = %s,
                not_severe = %s,
                severe = %s,
                death_total = %s,
                death_daily = %s,
                discharge_total = %s,
                discharge_daily = %s
            WHERE date = %s;
            """, (
                pcr_total,
                positive_up_total,
                positive_up_daily,
                positive,
                positive_yesterday,
                not_severe,
                severe,
                death_total,
                death_daily,
                discharge_total,
                discharge_daily,
                date,
            ))
